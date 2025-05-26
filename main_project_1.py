import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from fpdf import FPDF
import tkinter
import sqlite3


# Uncomment and comment print statements time and again to make program faster as easier to see
print("✅ All libraries are working fine!")

# A CSV (Comma-Separated Values) file is a plain text file that uses a specific structure to organize tabular data.
# The data present in CSV file are similar to excel but jsut are not seperated by lines or coloumns but they are sepereatee 
# by commas (,) . In general at top we write values that are to be examined seperated by commas and they the required vallues
# respectively below them also seperated by commas . 
import pandas as pd

df = pd.read_csv("students.csv")
print(df)

# df is generally used for dataFrameworks it is just a varialbe we could have used anything else

import pandas as pd
import matplotlib.pyplot as plt
from tkinter import Tk
# Uncomment from 28 to 35

print("✅ Pandas is working.")
print("✅ Matplotlib is working.")
print("✅ Tkinter is working.")

# Load CSV to check data again 

df = pd.read_csv("students.csv")

print("\n📊 Loaded Student Data:\n")
print(df.head())

##
# df.head() only shows the first 5 lines of the data , if want to see more data write "print(df)" or if want to see 
# data upto a fixed number of rows then write that number of rows in bracket of head like this : -> print(df.head(10))

# Step 2 – A: Performance Calculations

# 1. Calculate Total Marks
df["Total"] = df[["Maths", "Science", "English"]].sum(axis=1)

# 2. Calculate Percentage
df["Percentage"] = df["Total"] / 3  # since 3 subjects

# 3. Assign Grade Based on Percentage
def assign_grade(p):
    if p >= 90:
        return "A"
    elif p >= 75:
        return "B"
    elif p >= 60:
        return "C"
    elif p >= 40:
        return "D"
    else:
        return "F"

df["Grade"] = df["Percentage"].apply(assign_grade)

# 4. Print Updated Data
print("\n✅ Updated Student Performance Data:\n")
print(df[["Name", "Maths", "Science", "English", "Total", "Percentage", "Grade"]].head())

# The below shows how code from line 36 to 61 works

# For line 61 this how we write when we want to display only the things whcih we want to see and as
# above if we remove .head() or add something in the bracket how it behaves it is discussed earlier

"""

✅ Code Explanation: Step-by-Step
You already had this earlier:

python
Copy
Edit
df = pd.read_csv("students.csv")
This loads the student data into a DataFrame called df. Now, you added this new block:

🔹 Step 1: Calculate Total Marks
python
Copy
Edit
df["Total"] = df[["Maths", "Science", "English"]].sum(axis=1)
What it does:

Adds marks from Maths, Science, and English for each student.

Stores the result in a new column Total.

How it works:

df[["Maths", "Science", "English"]] selects those three columns.

.sum(axis=1) tells Pandas to add horizontally (row-wise).

🔹 Step 2: Calculate Percentage
python
Copy
Edit
df["Percentage"] = df["Total"] / 3
What it does:

Divides the total marks by 3 subjects.

Saves the result in a new column Percentage.

🔹 Step 3: Assign Grades
python
Copy
Edit
def assign_grade(p):
    if p >= 90:
        return "A"
    elif p >= 75:
        return "B"
    elif p >= 60:
        return "C"
    elif p >= 40:
        return "D"
    else:
        return "F"

df["Grade"] = df["Percentage"].apply(assign_grade)
What it does:

Defines a function assign_grade(p) that decides a grade based on the percentage p.

df["Percentage"].apply(assign_grade) applies this function to every student.

So:

If a student has 91% → gets grade A

If 65% → gets grade C

If 35% → gets grade F

🔹 Step 4: Print Clean Output
python
Copy
Edit
print("\n✅ Updated Student Performance Data:\n")
print(df[["Name", "Maths", "Science", "English", "Total", "Percentage", "Grade"]])
What it does:

Prints a nicely formatted table showing:

Name

Subject-wise marks

Total

Percentage

Grade

Only these columns are shown to keep it clean.

🧪 Example Output
mathematica
Copy
Edit
✅ Updated Student Performance Data:

      Name  Maths  Science  English  Total  Percentage Grade
0    Alice     78       85       90    253   84.33     B
1      Bob     55       60       58    173   57.67     D
2  Charlie     90       92       88    270   90.00     A
3    David     40       35       45    120   40.00     D

"""
# Step 2 – B: Identify Weakest Subject for Each Student

def find_weakest_subject(row):
    subjects = {"Maths": row["Maths"], "Science": row["Science"], "English": row["English"]}
    weakest = min(subjects, key=subjects.get)
    return weakest

df["Weakest Subject"] = df.apply(find_weakest_subject, axis=1)

# Print with Weakest Subject

print("\n📉 Student Data with Weakest Subjects:\n")
print(df[["Name", "Maths", "Science", "English", "Total", "Percentage", "Grade", "Weakest Subject"]])

# Below explains code from line 175 to 187
"""
🔍 Code Breakdown
def find_weakest_subject(row):
This function receives one row of the DataFrame at a time.

subjects = {"Maths": row["Maths"], "Science": row["Science"], "English": row["English"]}
We make a small dictionary to compare marks in all 3 subjects.

min(subjects, key=subjects.get)
Finds the subject name with the lowest marks.

df.apply(..., axis=1)
Applies this logic to each row (student) in the DataFrame.
"""
# Step 2 – C: Save updated data to a new CSV file
df.to_csv("analyzed_students.csv", index=False)

print("\n💾 Data saved successfully to 'analyzed_students.csv'")

# Below expalins code from line 210 to 213
"""
🧠 Code Explanation:
df.to_csv(...) → writes your full DataFrame to a CSV file.

"analyzed_students.csv" → is the name of the new file that will be created in the same folder.

index=False → skips writing the index column (0, 1, 2...).
"""
print(df.to_string(index=False)) # this prints df without index values on left 



# Step 3 – A: Bar Chart – Average Marks per Subject



subject_averages = {
    "Maths": df["Maths"].mean(),
    "Science": df["Science"].mean(),
    "English": df["English"].mean()
}

plt.figure(figsize=(8, 5))
plt.bar(subject_averages.keys(), subject_averages.values(), color=["skyblue", "lightgreen", "salmon"])
plt.title("📊 Average Marks per Subject")
plt.xlabel("Subjects")
plt.ylabel("Average Marks")
plt.ylim(0, 100)
plt.grid(axis='y', linestyle="--", alpha=0.5)
plt.tight_layout()
plt.show()

# Expalnation of code from 232 to 246 : ->
"""
🧠 Code Breakdown:
Line	Explanation
import matplotlib.pyplot as plt ->	Imports Matplotlib's plotting module
subject_averages = { ... } ->	Creates a dictionary with average marks
plt.bar(...) -> 	Draws a bar chart
plt.show() -> 	Displays the graph window
"""

# Step 3 – B: Pie Chart – Grade Distribution

grade_counts = df["Grade"].value_counts()

plt.figure(figsize=(7, 7))
plt.pie(
    grade_counts,
    labels=grade_counts.index,
    autopct='%1.1f%%',
    colors=["#4CAF50", "#2196F3", "#FFC107", "#FF5722", "#9C27B0"],
    startangle=140,
    wedgeprops={'edgecolor': 'black'}
)
plt.title("Grade Distribution")
plt.tight_layout()
plt.show()

# Below explains the code for lines 255 to 270 : ->
"""
🧠 Code Explanation:
Line	Meaning
grade_counts = df["Grade"].value_counts() ->	Counts how many students got each grade
plt.pie(...) ->	Creates the pie chart
autopct='%1.1f%%' -> 	Shows percentages (e.g. 25.0%)
startangle=140 ->	Rotates the pie chart for aesthetics
wedgeprops={'edgecolor': 'black'} ->	Adds borders to the slices for clarity



"""
# Step 3 – C: Line Graph – Individual Student Performance (with loop)

while True:
    roll_input = input("\n🔍 Enter Roll No. to visualize student performance (or type 'exit' to skip): ").strip()

    if roll_input.lower() == "exit":
        print("👋 Skipped student graph.")
        break

    if roll_input.isdigit() and int(roll_input) in df["Roll No"].values:
        student_row = df[df["Roll No"] == int(roll_input)].iloc[0]
        
        subjects = ["Maths", "Science", "English"]
        marks = [student_row["Maths"], student_row["Science"], student_row["English"]]

        plt.figure(figsize=(8, 5))
        plt.plot(subjects, marks, marker='o', linestyle='-', color='blue')
        plt.title(f"📈 Performance of {student_row['Name']} (Roll No: {roll_input})")
        plt.xlabel("Subjects")
        plt.ylabel("Marks")
        plt.ylim(0, 100)
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        plt.show()
        break  # exit loop after successful plot
    else:
        print("❌ Invalid Roll No. Please try again or type 'exit' to skip.")

# Explaantion of code from line 285 to 311 : ->
""" 
Lets the user exit gracefully by typing "exit".

🧠 Code Explanation:
Line	Purpose
input(...) ->	Asks user for a Roll No.
df[df["Roll No"] == int(roll_input)] ->	Filters row of that student
plt.plot(...) ->	Plots a line graph for that student’s marks
plt.title(...) ->	Personalizes the chart with student name and Roll No.
"""

# Step 4 – A: Export Final DataFrame to Excel
output_file = "student_report.xlsx"

try:
    df.to_excel(output_file, index=False)
    print(f"\nReport successfully saved as '{output_file}' in your project folder.")
except Exception as e:
    print(f"Failed to save report: {e}")

# Explanation of code from lines 325 to 332 : ->
""" 
🧠 Code Explanation:
Line	Meaning
df.to_excel(...) -> 	Saves the full DataFrame as an Excel file
index=False      ->	     Hides the left index column (0,1,2...)
try-except block -> 	Prevents the app from crashing if there's a save error
"""
from fpdf import FPDF

# Step 4 – B: Export Data to PDF
class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 18)
        self.cell(0, 10, "📄 Student Performance Report", ln=True, align="C")
        self.ln(5)

    def table(self, df):
        self.set_font("Arial", "B", 10)
        col_widths = [15, 30, 15, 15, 18, 20, 15, 20, 12, 30]

        # Table headers
        for i, col in enumerate(df.columns):
            self.cell(col_widths[i], 8, str(col), border=1)
        self.ln()

        # Table rows
        self.set_font("Arial", "B", 10)
        for _, row in df.iterrows():
            for i, col in enumerate(df.columns):
                self.cell(col_widths[i], 8, str(row[col]), border=1)
            self.ln()

pdf = FPDF()
pdf.add_page()
pdf.table(df)

pdf_file = "student_report.pdf"
try:
    pdf.output(pdf_file)
    print(f"PDF report saved as '{pdf_file}' in your folder.")
except Exception as e:
    print(f"Failed to save PDF: {e}")

# Expalnation of 344 to 379 : ->
"""
Try to understand later after reading file handling and class .
For now : 
 📂 Output:
A file named student_report.pdf will be created in your project folder with the full table.

📌 Notes:
You can adjust column widths and fonts as needed.

If the table is very large, we can break it into multiple pages
"""
# Now we will be using GUI ( Graphical User Interface) i.e. inbuilt python User Inerface designer to make input easier for 
# teachers or student to work with .

import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import pandas as pd

# GUI window setup
window = tk.Tk()
window.title("🎓 Student Performance Analyzer")
window.geometry("500x300")
window.configure(bg="#f0f0f0")

# Heading Label
title_label = tk.Label(window, text="📊 Student Performance Analyzer", font=("Arial", 16, "bold"), bg="#f0f0f0", fg="#333")
title_label.pack(pady=20)



# Button: Load Data
def load_data():
    try:
        df = pd.read_csv("students.csv")
        messagebox.showinfo("Success", f"Loaded {len(df)} student records!")
    except FileNotFoundError:
        messagebox.showerror("Error", "students.csv file not found!")

# --- Labels and Entries for Input Form ---
form_frame = tk.Frame(window, bg="#f0f0f0")
form_frame.pack(pady=10)

fields = ["Roll No", "Name", "Maths", "Science", "English", "Attendance"]
entries = {}

for i, field in enumerate(fields):
    label = tk.Label(form_frame, text=field + ":", bg="#f0f0f0", anchor='w', width=12)
    label.grid(row=i, column=0, padx=5, pady=2)
    
    entry = tk.Entry(form_frame, width=20)
    entry.grid(row=i, column=1, padx=5, pady=2)
    entries[field] = entry


load_btn = tk.Button(window, text="Load Student Data", command=load_data, width=25, height=2, bg="#4CAF50", fg="white")
load_btn.pack(pady=10)

# Exit Button
exit_btn = tk.Button(window, text="Exit", command=window.destroy, width=15, height=2, bg="#f44336", fg="white")
exit_btn.pack(pady=10)

def save_student():
    try:
        new_data = {field: entries[field].get().strip() for field in fields}
        
        # Validate numeric fields
        for subject in ["Maths", "Science", "English", "Attendance"]:
            new_data[subject] = int(new_data[subject])

        new_data["Total"] = new_data["Maths"] + new_data["Science"] + new_data["English"]
        new_data["Percentage"] = round(new_data["Total"] / 3, 2)

        # Assign Grade
        perc = new_data["Percentage"]
        if perc >= 90:
            grade = "A"
        elif perc >= 75:
            grade = "B"
        elif perc >= 60:
            grade = "C"
        elif perc >= 40:
            grade = "D"
        else:
            grade = "F"
        new_data["Grade"] = grade

        # Identify Weakest Subject
        subject_scores = {
            "Maths": new_data["Maths"],
            "Science": new_data["Science"],
            "English": new_data["English"]
        }
        new_data["Weakest Subject"] = min(subject_scores, key=subject_scores.get)

        df = pd.DataFrame([new_data])
        df.to_csv("students.csv", mode='a', index=False, header=False)
        messagebox.showinfo("Success", "Student added successfully!")

        for entry in entries.values():
            entry.delete(0, tk.END)

    except ValueError:
        messagebox.showerror("Error", "Please enter valid numeric values for marks and attendance.")
    except Exception as e:
        messagebox.showerror("Error", f"Unexpected error: {e}")

def view_students():
    try:
        df = pd.read_csv("students.csv")
        
        # Clear previous frame if exists
        for widget in result_frame.winfo_children():
            widget.destroy()
        
        # Display only first 10 rows if too many records
        rows_to_display = df.head(10)

        text = tk.Text(result_frame, height=15, width=80)
        text.pack(padx=10, pady=10)

        text.insert(tk.END, rows_to_display.to_string(index=False))

    except FileNotFoundError:
        messagebox.showerror("Error", "No student data found.")
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong: {e}")

save_btn = tk.Button(window, text="💾 Save Student", command=save_student, width=20, height=2, bg="#2196F3", fg="white")
save_btn.pack(pady=5)


view_btn = tk.Button(window, text="📖 View All Students", command=view_students, width=20, height=2, bg="#4CAF50", fg="white")
view_btn.pack(pady=5)



# Frame to display results (after buttons)
result_frame = tk.Frame(window, bg="#f0f0f0")
result_frame.pack(pady=10)





# Start the GUI loop
window.mainloop()

# Summarization of whta has  happened above : ->
"""
✅ Summary of Code Sections (Quick Visual Guide)
Section     	                       What it Does	                                           Placement
import	                       Imports required libraries	                                    Very top
window = tk.Tk()	              Sets up GUI window	                                      Right after imports
Input form (form_frame)	     Shows fields for student data input	                           Below window setup
save_btn	                    Button to trigger data saving	                                Below input form
save_student()	              Handles logic and saving to CSV	                              Near bottom, before mainloop
window.mainloop()	                Starts the GUI loop                                        	Last line of the file
"""